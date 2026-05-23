# ─────────────────────────────────────────────────────────────────────────────
# 0. 导入与配置
# ─────────────────────────────────────────────────────────────────────────────
import warnings, os
import numpy as np
import pandas as pd
from scipy import stats
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.rcParams.update({
    'font.family':        ['SimHei', 'Microsoft YaHei', 'Arial'],
    'axes.unicode_minus': False,
    'font.size':          10,
    'axes.spines.top':    False,
    'axes.spines.right':  False,
    'savefig.dpi':        300,
    'savefig.bbox':       'tight',
    'savefig.facecolor':  'white',
    'legend.frameon':     False,
})
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

warnings.filterwarnings('ignore')

os.makedirs('论文图表', exist_ok=True)

# 全局颜色常量（图表使用）
C_DARK  = '#2B3A52'
C_MID   = '#5B7FA6'
C_LIGHT = '#A8BFCF'
C_GRAY  = '#7A7A7A'
C_LGRAY = '#BABABA'
# ─────────────────────────────────────────────────────────────────────────────
# 1. 数据加载（四文件合并）
# ─────────────────────────────────────────────────────────────────────────────

# ══ 修改这里：四个 Excel 文件的路径 ══════════════════════════════════════════
PATH_BASE_STU  = r'基线学生.xlsx'
PATH_BASE_PAR  = r'基线家长.xlsx'
PATH_FOLLOW_STU = r'追加学生数据.xlsx'
PATH_FOLLOW_PAR = r'追加家长数据.xlsx'
# ═════════════════════════════════════════════════════════════════════════════

def load_data(_=None) -> pd.DataFrame:
    """
    按你的原始代码合并四个文件，去重后返回宽表。
    如果追加文件不存在则只用基线数据。
    """
    baseline_stu = pd.read_excel(PATH_BASE_STU)
    baseline_par = pd.read_excel(PATH_BASE_PAR)
    baseline_data = pd.merge(baseline_stu, baseline_par, on='ids', how='inner')
    print(f"基线数据：{len(baseline_data):,} 行")

    try:
        follow_stu  = pd.read_excel(PATH_FOLLOW_STU)
        follow_par  = pd.read_excel(PATH_FOLLOW_PAR)
        follow_data = pd.merge(follow_stu, follow_par, on='ids', how='inner')
        data = (pd.concat([baseline_data, follow_data], ignore_index=True)
                  .drop_duplicates(subset='ids'))
        print(f"追加数据：{len(follow_data):,} 行")
    except FileNotFoundError:
        print("  ⚠ 追加数据文件未找到，仅使用基线数据")
        data = baseline_data

    print(f"合并去重后：{len(data):,} 行 × {data.shape[1]} 列")
    print(f"  列名预览：{list(data.columns[:15])}")
    return data


# ─────────────────────────────────────────────────────────────────────────────
# 2. 变量构建
# ─────────────────────────────────────────────────────────────────────────────

# ══ 如果列名与默认值不同，在这里改 ══════════════════════════════════════════
COL_MAP = {
    # 同胞变量 ── 诊断确认存在 b0201-b0204 ✅
    'older_brother':   'b0201',
    'younger_brother': 'b0202',
    'older_sister':    'b0203',
    'younger_sister':  'b0204',
    # 成绩 ── 诊断确认 tr_chn/tr_mat/tr_eng ✅
    'chinese_raw':  'tr_chn',
    'math_raw':     'tr_mat',
    'english_raw':  'tr_eng',
    # 认知能力 ── 诊断确认 stcog ✅
    'cognitive':    'stcog',
    # 控制变量
    'gender':   'stsex',     # ✅ 已确认
    'hukou':    'ba18',      # 家长问卷户籍（1=农业, 2=非农业/城镇）
    'fa_edu':   'a14',       # 学生问卷父亲教育程度
    'mo_edu':   'a15',       # 学生问卷母亲教育程度
}
# ═════════════════════════════════════════════════════════════════════════════


def diagnose_columns(df: pd.DataFrame):
    """
    打印所有与关键变量相关的列名，帮助确认 COL_MAP 是否正确。
    运行一次后根据输出修改 COL_MAP，之后不再需要调用此函数。
    """
    print("\n" + "="*60)
    print("【列名诊断】请根据以下输出核对 COL_MAP")
    print("="*60)
    cols = list(df.columns)

    def show(title, patterns):
        matched = [c for c in cols if any(p.lower() in c.lower() for p in patterns)]
        print(f"\n▶ {title}（找到 {len(matched)} 列）：")
        if matched:
            print("  ", matched[:30])
        else:
            print("   （未找到，请手动搜索）")

    show("成绩相关 (tr_ / chn/mat/eng/score)",
         ['tr_', 'chn', 'mat', 'eng', 'score', 'std'])
    show("认知能力 (cog / stcog / iq)",
         ['cog', 'stcog', 'iq', 'cognitive'])
    show("同胞兄弟姐妹 (b02 / sibling / bro / sis)",
         ['b020', 'sibling', 'bro', 'sis', 'a02'])
    show("性别 (sex / gender)",
         ['sex', 'gender'])
    show("户口 (hukou / a18 / regist / urban)",
         ['hukou', 'a18', 'regist', 'urban', 'a17', 'a19'])
    show("父亲教育 (a14 / faedu / bfaedu)",
         ['a14', 'faedu', 'bfaedu', 'fa_edu', 'father'])
    show("母亲教育 (a15 / moedu / bmoedu)",
         ['a15', 'moedu', 'bmoedu', 'mo_edu', 'mother'])
    print("\n" + "="*60 + "\n")


def _find_col(df: pd.DataFrame, *candidates) -> str | None:
    """在 df 中找第一个存在的候选列名，都不存在返回 None"""
    for c in candidates:
        if c and c in df.columns:
            return c
    return None


def build_variables(raw: pd.DataFrame) -> pd.DataFrame:
    """
    严格按照用户实际代码（步骤3-9）构建所有分析变量。
    列名以实际数据为准：stfedu/stmedu/sthktype/b01等。
    """
    from sklearn.preprocessing import StandardScaler

    df = raw.copy()
    print(f"\n【列名确认】共 {df.shape[1]} 列，前40列：")
    print(list(df.columns[:40]))

    # ── 步骤3：剔除独生子女 ───────────────────────────────────────────────────
    total_raw = len(df)
    only_child = (df['b01'] == 1).sum()
    df = df[df['b01'] == 2].copy()
    print(f"\n原始样本量：{total_raw:,}")
    print(f"独生子女（b01=1）：{only_child:,}")
    print(f"剔除独生子女后：{len(df):,}")

    # ── 步骤4：同胞数量 & 剔除矛盾样本 ──────────────────────────────────────
    df['total_siblings'] = df[['b0201','b0202','b0203','b0204']].sum(axis=1)
    df = df[df['total_siblings'] >= 1].copy()
    print(f"剔除矛盾样本后：{len(df):,}")

    # ── 步骤5：核心处理变量 ───────────────────────────────────────────────────
    df['has_sister']   = np.where(df['b0203'] >= 1, 1, 0)
    df['has_brother']  = np.where(df['b0201'] >= 1, 1, 0)
    df['is_multi']     = (df['total_siblings'] >= 2).astype(int)
    df['is_two_child'] = (df['total_siblings'] == 1).astype(int)
    df['sister_multi'] = df['has_sister'] * df['is_multi']
    df['is_secondborn']= ((df['b0201'] + df['b0203']) >= 1).astype(int)

    # 二孩家庭：仅有一个年长姐姐，无其他同胞
    df['two_child_only_sister'] = (
        (df['total_siblings'] == 1) & (df['b0203'] == 1) &
        (df['b0201'] == 0) & (df['b0202'] == 0) & (df['b0204'] == 0)
    ).astype(int)

    # 多孩家庭有姐姐
    df['multi_with_sister']  = ((df['has_sister'] == 1) & (df['total_siblings'] >= 2)).astype(int)
    df['no_sister']          = (df['has_sister'] == 0).astype(int)

    # 多孩家庭内部细分（大姐 vs 有哥有姐）
    df['only_sister_multi']  = ((df['multi_with_sister'] == 1) & (df['b0201'] == 0)).astype(int)
    df['sister_and_brother'] = ((df['multi_with_sister'] == 1) & (df['b0201'] >= 1)).astype(int)

    # 有年幼妹妹（机制分析用）
    df['has_younger_sister'] = (df['b0204'] >= 1).astype(int)

    # ── 步骤6：控制变量 ───────────────────────────────────────────────────────
    df['par_edu'] = (pd.to_numeric(df['stfedu'], errors='coerce') +
                     pd.to_numeric(df['stmedu'], errors='coerce')) / 2
    df['hukou']   = np.where(df['sthktype'] == 1, 1, 0)   # 1=城镇，0=农村
    df['gender']  = df['stsex']
    print(f"  ✅ 性别变量 stsex  |  男生占比：{df['gender'].mean():.2%}")
    print(f"  ✅ 户口变量 sthktype  |  城镇占比：{df['hukou'].mean():.2%}")
    print(f"  ✅ 父母教育 stfedu/stmedu  |  均值：{df['par_edu'].mean():.2f}")

    # ── 步骤7：标准化成绩 ─────────────────────────────────────────────────────
    subjects = ['tr_chn', 'tr_mat', 'tr_eng']
    scaler = StandardScaler()

    # 转为数值类型
    for sub in subjects:
        df[sub] = pd.to_numeric(df[sub], errors='coerce')

    # 标准化
    scaled = scaler.fit_transform(df[subjects])
    df[['std_chn', 'std_mat', 'std_eng']] = scaled

    # ✅ 新增：只保留三科成绩都齐全的样本
    before_drop = len(df)
    df = df.dropna(subset=['std_chn', 'std_mat', 'std_eng']).copy()
    print(f"  标准化后删除成绩缺失样本：{before_drop - len(df):,} 行，剩余 {len(df):,}")

    # 计算 GPA（此时三科都不缺失）
    df['gpa'] = df[['std_chn', 'std_mat', 'std_eng']].mean(axis=1)

    # 认知能力（stcog 已标准化，直接用）
    if 'stcog' in df.columns:
        df['cognitive'] = pd.to_numeric(df['stcog'], errors='coerce')
        print(f"  ✅ 认知能力变量 stcog")
    else:
        df['cognitive'] = np.nan
        print("  ⚠ 未找到认知能力变量 stcog")

    # 兼容旧列名（回归函数中使用）
    df['chinese_std'] = df['std_chn']
    df['math_std'] = df['std_mat']
    df['english_std'] = df['std_eng']

    # ── 步骤8：清理缺失值 ─────────────────────────────────────────────────────
    before_final = len(df)
    df = df.dropna(subset=['has_sister', 'gender', 'par_edu', 'hukou', 'total_siblings']).copy()
    print(f"  清理控制变量缺失：{before_final - len(df):,} 行")
    print(f"\n清理缺失值后最终样本量：{len(df):,}")
    print(f"  has_sister 均值：{df['has_sister'].mean():.3f}（有姐姐占比）")

    return df


    # ── 步骤8：清理缺失值 ─────────────────────────────────────────────────────
    df = df.dropna(subset=['gpa','has_sister','gender','par_edu','hukou','total_siblings']).copy()
    print(f"\n清理缺失值后最终样本量：{len(df):,}")
    print(f"  has_sister 均值：{df['has_sister'].mean():.3f}（有姐姐占比）")

    return df


# ─────────────────────────────────────────────────────────────────────────────
# 3. 样本构建（三套样本，严格按实际分析逻辑）
# ─────────────────────────────────────────────────────────────────────────────

def build_samples(df: pd.DataFrame):
    """
    三套分析样本：
      A. 宽二孩样本（samp_broad） —— 二孩家庭（total_siblings==1），用于机制分析、灵敏度对比
      B. 异质性样本               —— 全部非独生子女（total_siblings>=1），用于家庭规模异质性
      C. 主回归样本（samp_main）  —— 二孩次子女严格对比（有年长姐姐 vs 有年长哥哥）
                                      对照组：有年长哥哥且无年长姐姐的次子女
                                      处置组：有年长姐姐且无年长哥哥的次子女
    注意：本研究以样本C为主回归样本，样本A仅用于机制分析与灵敏度对比。
    """
    CORE_VARS = ['has_sister','gender','par_edu','hukou','gpa',
                 'std_chn','std_mat','std_eng']

    # ── 样本 B：全部非独生子女（步骤3-8 后的 df 已剔除独生子女和矛盾样本）────
    samp_B = df.dropna(subset=['is_multi']).copy()
    print(f"\n样本B（异质性，二孩+多孩）：N = {len(samp_B):,}")
    print(f"  二孩家庭：N = {(samp_B['is_two_child']==1).sum():,}")
    print(f"  多孩家庭：N = {(samp_B['is_multi']==1).sum():,}")

    # ── 样本 A：二孩家庭（total_siblings==1）─────────────────────────────────
    samp_A = samp_B[samp_B['is_two_child'] == 1].copy()
    print(f"\n样本A（主回归，二孩）：N = {len(samp_A):,}")
    print(f"  有姐姐：N = {samp_A['two_child_only_sister'].sum():,}（{samp_A['two_child_only_sister'].mean():.1%}）")

    # ── 样本 C：二孩次子女严格对比（有年长姐姐 或 有年长哥哥）────────────────
    samp_C = samp_A[
        samp_A['two_child_only_sister'] | ((samp_A['b0201']>=1)&(samp_A['b0203']==0))   # 有年长同胞 = 次子女
    ].copy()
    print(f"\n样本C（严格对比，二孩次子女）：N = {len(samp_C):,}")
    print(f"  有姐姐：{samp_C['two_child_only_sister'].sum():,}  |  有哥哥无姐姐：{((samp_C['b0201']>=1)&(samp_C['b0203']==0)).sum():,}")

    # ── 样本 D：多孩次子女严格对比（有年长姐姐 OR 有年长哥哥）──────────────
    # 对应论文表12"多孩"行：多孩家庭中、是次子女、且有年长姐姐或年长哥哥
    samp_C_multi = samp_B[
        (samp_B['is_multi'] == 1) &
        ((samp_B['b0203'] >= 1) | (samp_B['b0201'] >= 1)) &
        (samp_B['is_secondborn'] == 1)
    ].copy()
    print(f"\n样本D（多孩严格对比，次子女）：N = {len(samp_C_multi):,}")
    if len(samp_C_multi) > 0:
        sister_D  = int((samp_C_multi['b0203'] >= 1).sum())
        brother_D = int(((samp_C_multi['b0201'] >= 1) & (samp_C_multi['b0203'] == 0)).sum())
        print(f"  有年长姐姐：{sister_D:,}  |  有年长哥哥（无姐）：{brother_D:,}")

    # ── 样本统计汇总（对应论文表1）───────────────────────────────────────────
    two   = samp_B[samp_B['is_two_child'] == 1]
    multi = samp_B[samp_B['is_multi']     == 1]
    sister_C = (samp_C['b0203'] >= 1).sum()
    brother_C = ((samp_C['b0201'] >= 1) & (samp_C['b0203'] == 0)).sum()
    print(f"\n【样本统计汇总（表1）】")
    print(f"  全部有效分析样本：{len(samp_B):,}")
    print(f"  二孩家庭总量：    {len(two):,}")
    print(f"  二孩有姐姐：      {two['two_child_only_sister'].sum():,}（{two['two_child_only_sister'].mean():.2%}）")
    print(f"  多孩家庭总量：    {len(multi):,}")
    print(f"  多孩有姐姐：      {multi['has_sister'].sum():,}（{multi['has_sister'].mean():.2%}）")
    print(f"\n【主回归样本（样本C，严格对比）】N = {len(samp_C):,}")
    print(f"  处置组（有年长姐姐）：{sister_C:,}（{sister_C/len(samp_C):.2%}）")
    print(f"  对照组（有年长哥哥）：{brother_C:,}（{brother_C/len(samp_C):.2%}）")
    print(f"  对照组选择说明：次子女中仅有年长哥哥者，出生顺序与年龄差结构与处置组对称，")
    print(f"  排除'是否有同胞'的选择效应，has_sister 系数识别纯粹同胞性别效应。")

    return samp_A, samp_B, samp_C, samp_C_multi


# ─────────────────────────────────────────────────────────────────────────────
# 4. 回归工具函数
# ─────────────────────────────────────────────────────────────────────────────

CTRL = 'gender + par_edu + hukou'

def ols(df, y, x_list, robust='HC3'):
    """运行 OLS，返回 results 对象"""
    X = sm.add_constant(df[x_list].astype(float))
    Y = df[y].astype(float)
    mask = X.notna().all(axis=1) & Y.notna()
    m = sm.OLS(Y[mask], X[mask]).fit(cov_type=robust)
    return m

def extract_row(m, varname, label=None):
    """从回归结果提取一行：系数、标准误、p值、显著性"""
    if varname not in m.params.index:
        return {'变量': label or varname, '系数': '—', '标准误': '—',
                'p值': '—', '显著性': '—', '95%CI下界': '—', '95%CI上界': '—'}
    b  = m.params[varname]
    se = m.bse[varname]
    p  = m.pvalues[varname]
    lo, hi = m.conf_int().loc[varname]
    sig = '***' if p < 0.01 else ('**' if p < 0.05 else ('*' if p < 0.10 else ''))
    return {'变量': label or varname,
            '系数': round(b, 4), '标准误': round(se, 4),
            'p值':  round(p, 4), '显著性': sig,
            '95%CI下界': round(lo, 4), '95%CI上界': round(hi, 4)}

def model_summary(m, key_vars, extra=None):
    """返回回归结果的 DataFrame（论文格式）"""
    rows = [extract_row(m, v) for v in key_vars]
    rows.append({'变量': '样本量', '系数': int(m.nobs), '标准误': '',
                 'p值': '', '显著性': '', '95%CI下界': '', '95%CI上界': ''})
    rows.append({'变量': 'R²', '系数': round(m.rsquared, 4), '标准误': '',
                 'p值': '', '显著性': '', '95%CI下界': '', '95%CI上界': ''})
    if extra:
        for k, v in extra.items():
            rows.append({'变量': k, '系数': v, '标准误': '',
                         'p值': '', '显著性': '', '95%CI下界': '', '95%CI上界': ''})
    return pd.DataFrame(rows)


def bootstrap_coef(df, y, x_list, target_var, n_boot=500, seed=42):
    """Bootstrap 估计目标系数的均值、SE 和 p 值"""
    np.random.seed(seed)
    coefs = []
    for _ in range(n_boot):
        sample = df.sample(len(df), replace=True)
        try:
            m = ols(sample, y, x_list, robust='HC3')
            if target_var in m.params.index:
                coefs.append(m.params[target_var])
        except Exception:
            continue
    coefs = np.array(coefs)
    beta  = np.mean(coefs)
    se    = np.std(coefs, ddof=1)
    # 两侧 p 值（基于 t 分布）
    t_stat = beta / se if se > 0 else 0
    p_val  = 2 * (1 - stats.t.cdf(abs(t_stat), df=len(df)-len(x_list)-1))
    ci_lo  = np.percentile(coefs, 2.5)
    ci_hi  = np.percentile(coefs, 97.5)
    return beta, se, p_val, ci_lo, ci_hi


# ─────────────────────────────────────────────────────────────────────────────
# 5. 各节分析函数
# ─────────────────────────────────────────────────────────────────────────────

# ── 5.1 描述性统计（表1、表2、表3）──────────────────────────────────────────

def analysis_descriptive(samp_main, samp_B, samp_broad=None, samp_C_multi=None):
    """
    表1：样本构成（含主样本C详情）
    表2：描述性统计（主回归样本 = 严格对比样本C）
    表3：核心变量相关矩阵
    表12desc：多孩严格对比描述性统计（有年长姐姐 vs 有年长哥哥）
    samp_main    = 主回归样本（样本C：二孩次子女严格对比）
    samp_B       = 全部非独生子女（家庭规模异质性分析用）
    samp_broad   = 宽二孩样本（样本A），用于对比展示
    samp_C_multi = 多孩严格对比样本（样本D），用于表12多孩行
    """
    results = {}

    # ── 表1：样本构成 ─────────────────────────────────────────────────────────
    two = samp_B[samp_B['is_two_child'] == 1]
    mul = samp_B[samp_B['is_multi'] == 1]
    sister_C  = int((samp_main['b0203'] >= 1).sum())
    brother_C = int(((samp_main['b0201'] >= 1) & (samp_main['b0203'] == 0)).sum())
    t1_data = {
        '统计项目': [
            '全部有效分析样本量',
            '独生子女（已剔除）',
            '二孩家庭总样本量（宽样本A）',
            '二孩家庭中有姐姐',
            '多孩家庭总样本量',
            '多孩家庭中有姐姐',
            '── 主回归样本（严格对比，样本C）──',
            '  处置组：有年长姐姐的次子女',
            '  对照组：有年长哥哥的次子女',
        ],
        '样本量': [
            len(samp_B),
            '—',
            len(two),
            int(two['two_child_only_sister'].sum()),
            len(mul),
            int(mul['has_sister'].sum()),
            len(samp_main),
            sister_C,
            brother_C,
        ],
        '占比(%)': [
            '100',
            '—',
            f'{100*len(two)/len(samp_B):.2f}',
            f'{100*two["has_sister"].mean():.2f}',
            f'{100*len(mul)/len(samp_B):.2f}',
            f'{100*mul["has_sister"].mean():.2f}',
            '—',
            f'{100*sister_C/len(samp_main):.2f}',
            f'{100*brother_C/len(samp_main):.2f}',
        ],
    }
    results['表1_样本构成'] = pd.DataFrame(t1_data)
    print("✅ 表1 完成")

    # ── 表2：描述性统计（主回归样本 = 样本C）──────────────────────────────────
    vars_desc = {
        'gpa':         '标准化GPA',
        'chinese_std': '语文标准化成绩',
        'math_std':    '数学标准化成绩',
        'english_std': '英语标准化成绩',
        'has_sister':  '有年长姐姐（=1，对照=有年长哥哥）',
        'gender':      '学生性别（男=1）',
        'par_edu':     '父母教育程度（等级均值）',
        'hukou':       '城镇户口（=1）',
    }
    rows = []
    for var, label in vars_desc.items():
        s = samp_main[var].dropna()
        rows.append({
            '变量': label, '均值': round(s.mean(),4),
            '标准差': round(s.std(),4), '最小值': round(s.min(),4),
            '最大值': round(s.max(),4), '观测数': int(s.count()),
        })
    results['表2_描述性统计'] = pd.DataFrame(rows)

    # 按有无姐姐分组均值对比（严格对比：有姐 vs 有哥）
    grp = []
    for var, label in vars_desc.items():
        for g, gname in [(0,'对照组（有年长哥哥）'),(1,'处置组（有年长姐姐）')]:
            s = samp_main[samp_main['has_sister']==g][var].dropna()
            grp.append({'变量':label, '分组':gname,
                        '均值':round(s.mean(),4), '样本量':int(s.count())})
    results['表2b_分组描述统计'] = pd.DataFrame(grp)

    # 宽样本对比（如果提供）
    if samp_broad is not None:
        rows_b = []
        for var, label in vars_desc.items():
            s = samp_broad[var].dropna()
            rows_b.append({
                '变量': label, '均值': round(s.mean(),4),
                '标准差': round(s.std(),4), '观测数': int(s.count()),
            })
        results['表2c_宽样本描述（参考）'] = pd.DataFrame(rows_b)
    print("✅ 表2 完成")

    # ── 表3：相关矩阵 ─────────────────────────────────────────────────────────
    corr_vars = ['gpa','has_sister','gender','par_edu','hukou']
    corr_labels = ['GPA','有年长姐姐','学生性别','父母教育','城镇户口']
    C = samp_main[corr_vars].corr()
    C.index   = corr_labels
    C.columns = corr_labels
    # 计算 p 值矩阵
    p_mat = pd.DataFrame(index=corr_labels, columns=corr_labels)
    for i, v1 in enumerate(corr_vars):
        for j, v2 in enumerate(corr_vars):
            if i == j:
                p_mat.iloc[i,j] = ''
            else:
                _, p = stats.pearsonr(
                    samp_main[v1].dropna().align(samp_main[v2].dropna(), join='inner')[0],
                    samp_main[v1].dropna().align(samp_main[v2].dropna(), join='inner')[1],
                )
                star = '***' if p<0.001 else ('**' if p<0.01 else ('*' if p<0.05 else ''))
                p_mat.iloc[i,j] = f'{C.iloc[i,j]:.3f}{star}'
    results['表3_相关矩阵'] = p_mat
    print("✅ 表3 完成")

    # ── 表12辅助：多孩严格对比分组描述性统计 ─────────────────────────────────
    if samp_C_multi is not None and len(samp_C_multi) > 0:
        subj_vars = [
            ('gpa',         'GPA'),
            ('chinese_std', '语文'),
            ('math_std',    '数学'),
            ('english_std', '英语'),
        ]
        multi_rows = []
        for g, gname in [(1,'有姐姐（多孩处置组）'), (0,'有哥哥（多孩对照组）')]:
            sub = samp_C_multi[samp_C_multi['has_sister'] == g]
            row = {'组别': gname, 'N': int(len(sub))}
            for var, label in subj_vars:
                row[label + '均值'] = round(sub[var].dropna().mean(), 4)
            multi_rows.append(row)
        # 均值差行
        sis_sub = samp_C_multi[samp_C_multi['has_sister'] == 1]
        bro_sub = samp_C_multi[samp_C_multi['has_sister'] == 0]
        diff_row = {'组别': '均值差（姐−哥）', 'N': '—'}
        for var, label in subj_vars:
            d = sis_sub[var].dropna().mean() - bro_sub[var].dropna().mean()
            diff_row[label + '均值'] = round(d, 4)
        multi_rows.append(diff_row)
        results['表12_多孩严格对比描述统计'] = pd.DataFrame(multi_rows)

        # 多孩t检验（均值差显著性）
        ttest_rows = []
        for var, label in subj_vars:
            x1 = sis_sub[var].dropna()
            x0 = bro_sub[var].dropna()
            t, p = stats.ttest_ind(x1, x0, equal_var=False)
            sig = '***' if p<0.01 else ('**' if p<0.05 else ('*' if p<0.10 else 'ns'))
            ttest_rows.append({'科目': label, 't统计量': round(t,4),
                               'p值': round(p,4), '显著性': sig,
                               '有姐姐均值': round(x1.mean(),4),
                               '有哥哥均值': round(x0.mean(),4),
                               '均值差': round(x1.mean()-x0.mean(),4)})
        results['表12_多孩均值差t检验'] = pd.DataFrame(ttest_rows)
        print("✅ 表12多孩描述统计 完成")

    # 二孩严格对比 t 检验（补全表10均值差显著性）
    subj_vars2 = [('gpa','GPA'),('chinese_std','语文'),('math_std','数学'),('english_std','英语')]
    ttest2_rows = []
    for var, label in subj_vars2:
        x1 = samp_main[samp_main['has_sister']==1][var].dropna()
        x0 = samp_main[samp_main['has_sister']==0][var].dropna()
        t, p = stats.ttest_ind(x1, x0, equal_var=False)
        sig = '***' if p<0.01 else ('**' if p<0.05 else ('*' if p<0.10 else 'ns'))
        ttest2_rows.append({'科目': label, 't统计量': round(t,4),
                            'p值': round(p,4), '显著性': sig,
                            '有姐姐均值': round(x1.mean(),4),
                            '有哥哥均值': round(x0.mean(),4),
                            '均值差': round(x1.mean()-x0.mean(),4)})
    results['表10_二孩均值差t检验'] = pd.DataFrame(ttest2_rows)
    print("✅ 表10二孩t检验 完成")

    return results


# ── 5.2 主回归：三模型逐步回归（表4）────────────────────────────────────────

def analysis_main_regression(samp_main):
    """
    表4：姐姐效应对学业成绩的回归结果（模型1/2/3）
    主回归样本 = 严格对比样本C（有年长姐姐 vs 有年长哥哥的二孩次子女）
    has_sister = 1 → 有年长姐姐；has_sister = 0 → 有年长哥哥（对照组）
    """
    results = {}

    specs = [
        ('模型1_无控制', ['has_sister']),
        ('模型2_人口学控制', ['has_sister','gender']),
        ('模型3_全控制', ['has_sister','gender','par_edu','hukou']),
    ]
    all_rows = []
    for name, xvars in specs:
        m = ols(samp_main, 'gpa', xvars)
        for v in xvars + (['const'] if True else []):
            r = extract_row(m, v)
            r['模型'] = name
            all_rows.append(r)
        all_rows.append({'变量':'样本量', '模型':name, '系数':int(m.nobs),
                         '标准误':'', 'p值':'', '显著性':'', '95%CI下界':'', '95%CI上界':''})
        all_rows.append({'变量':'R²', '模型':name, '系数':round(m.rsquared,4),
                         '标准误':'', 'p值':'', '显著性':'', '95%CI下界':'', '95%CI上界':''})
        all_rows.append({'变量':'调整R²', '模型':name, '系数':round(m.rsquared_adj,4),
                         '标准误':'', 'p值':'', '显著性':'', '95%CI下界':'', '95%CI上界':''})
    df_out = pd.DataFrame(all_rows)
    # 宽格式：每模型一列
    wide = df_out.pivot_table(index='变量', columns='模型', values='系数',
                              aggfunc='first').reset_index()
    results['表4_主回归（严格对比样本C）'] = df_out
    results['表4b_主回归_宽格式'] = wide
    print("✅ 表4 完成（主回归样本：严格对比样本C）")
    return results


# ── 5.3 分科目回归（表5）────────────────────────────────────────────────────

def analysis_subjects(samp_main):
    """
    表5：分科目姐姐效应——完整回归系数（严格对比样本C）
    输出三张表：
      表5_完整宽格式  → 论文直接使用（行=变量，列=科目，单元格=系数(SE)）
      表5_完整长格式  → 每科目完整系数+SE+p+CI
      表5_宽格式_数值 → 纯数值（便于复制）
    """
    results = {}
    xvars = ['has_sister', 'gender', 'par_edu', 'hukou']
    xvar_labels = {
        'has_sister': '有姐姐（年长）',
        'gender':     '学生性别',
        'par_edu':    '父母平均教育年限',
        'hukou':      '城镇户口',
        'const':      '常数项',
    }
    # 论文列顺序：语文 | 数学 | 英语 | GPA
    subjects = [
        ('chinese_std', '语文'),
        ('math_std',    '数学'),
        ('english_std', '英语'),
        ('gpa',         'GPA'),
    ]

    # ── 运行所有科目回归 ──────────────────────────────────────────────────────
    models = {}
    for y, label in subjects:
        models[label] = ols(samp_main, y, xvars)

    # ── 宽格式：每格 "β***(SE)" ───────────────────────────────────────────────
    wide_rows = []
    for v in xvars + ['const']:
        coef_row = {'变量': xvar_labels.get(v, v)}
        for _, label in subjects:
            m = models[label]
            if v in m.params.index:
                b   = m.params[v]
                se  = m.bse[v]
                p   = m.pvalues[v]
                sig = '***' if p<0.01 else ('**' if p<0.05 else ('*' if p<0.10 else ''))
                coef_row[label] = f'{b:.4f}{sig} ({se:.4f})'
            else:
                coef_row[label] = '—'
        wide_rows.append(coef_row)

    # 样本量 / R² / 调整R²
    for stat_name, getter in [
        ('样本量',  lambda m: int(m.nobs)),
        ('R²',     lambda m: round(m.rsquared, 4)),
        ('调整R²', lambda m: round(m.rsquared_adj, 4)),
    ]:
        row = {'变量': stat_name}
        for _, label in subjects:
            row[label] = getter(models[label])
        wide_rows.append(row)

    results['表5_分科目_完整宽格式'] = pd.DataFrame(wide_rows)

    # ── 纯数值宽格式（便于核对）──────────────────────────────────────────────
    num_rows = []
    for v in xvars + ['const']:
        for stat, col_sfx in [('系数','_β'), ('标准误','_SE'), ('p值','_p')]:
            row = {'变量': xvar_labels.get(v, v), '统计量': stat}
            for _, label in subjects:
                m = models[label]
                if v in m.params.index:
                    if stat == '系数':
                        row[label] = round(m.params[v], 4)
                    elif stat == '标准误':
                        row[label] = round(m.bse[v], 4)
                    else:
                        row[label] = round(m.pvalues[v], 4)
                else:
                    row[label] = '—'
            num_rows.append(row)
    # 样本量 / R²
    for stat_name, getter in [('样本量', lambda m: int(m.nobs)), ('R²', lambda m: round(m.rsquared,4))]:
        row = {'变量': stat_name, '统计量': ''}
        for _, label in subjects:
            row[label] = getter(models[label])
        num_rows.append(row)
    results['表5_分科目_数值明细'] = pd.DataFrame(num_rows)

    # ── 长格式：每科目完整系数行（含CI）────────────────────────────────────
    long_rows = []
    for y, label in subjects:
        m = models[label]
        for v in xvars + ['const']:
            r = extract_row(m, v, label=xvar_labels.get(v, v))
            r['科目'] = label
            long_rows.append(r)
        long_rows.append({'变量': '样本量', '系数': int(m.nobs), '科目': label,
                          '标准误':'','p值':'','显著性':'','95%CI下界':'','95%CI上界':''})
        long_rows.append({'变量': 'R²', '系数': round(m.rsquared,4), '科目': label,
                          '标准误':'','p值':'','显著性':'','95%CI下界':'','95%CI上界':''})
        long_rows.append({'变量': '调整R²', '系数': round(m.rsquared_adj,4), '科目': label,
                          '标准误':'','p值':'','显著性':'','95%CI下界':'','95%CI上界':''})
    results['表5_分科目_完整长格式'] = pd.DataFrame(long_rows)

    print("✅ 表5 完成（完整输出：宽格式/数值明细/长格式，共3张表）")
    # 返回 models 时顺序与原逻辑兼容（gen_figures 使用 GPA/语文/数学/英语 key）
    models_compat = {label: models[label] for _, label in [('gpa','GPA'),('chinese_std','语文'),
                                                            ('math_std','数学'),('english_std','英语')]}
    return results, models_compat


# ── 5.4 性别异质性（表6）────────────────────────────────────────────────────

def analysis_gender(samp_main):
    """
    表6a：有姐姐×性别 交互项模型（严格对比样本C）
    表6b：分性别子样本回归（严格对比样本C）
    对照组在两个子样本中均为有年长哥哥的次子女
    """
    results = {}
    subjects = [('gpa','GPA'), ('chinese_std','语文'),
                ('math_std','数学'), ('english_std','英语')]
    interact_rows = []
    for y, label in subjects:
        samp = samp_main.copy()
        samp['sister_girl'] = samp['has_sister'] * (1 - samp['gender'])
        m = ols(samp, y, ['has_sister','gender','sister_girl','par_edu','hukou'])
        for v, lbl in [('has_sister',f'{label}_有姐姐（男生基准）'),
                       ('gender',    f'{label}_性别'),
                       ('sister_girl',f'{label}_姐姐×女生')]:
            r = extract_row(m, v, label=lbl)
            r['科目'] = label
            interact_rows.append(r)
        interact_rows.append({'变量':f'{label}_样本量','系数':int(m.nobs),
                              '科目':label,'标准误':'','p值':'','显著性':'',
                              '95%CI下界':'','95%CI上界':''})
    results['表6a_性别交互项（样本C）'] = pd.DataFrame(interact_rows)

    # 分性别子样本
    sub_rows = []
    for g, gname in [(1,'男生（弟弟）'),(0,'女生（妹妹）')]:
        sub = samp_main[samp_main['gender'] == g].copy()
        for y, label in subjects:
            m = ols(sub, y, ['has_sister','par_edu','hukou'])
            r = extract_row(m, 'has_sister', label=f'{gname}_{label}')
            r['性别'] = gname
            r['科目'] = label
            r['样本量'] = int(m.nobs)
            sub_rows.append(r)
    results['表6b_分性别子样本（样本C）'] = pd.DataFrame(sub_rows)
    print("✅ 表6 完成（主回归样本：严格对比样本C）")
    return results


# ── 5.5 家庭规模异质性：分样本回归（表7）────────────────────────────────────

def analysis_family_size_subsample(samp_B):
    """表7：二孩 vs 多孩 分样本回归（四科目×两组）"""
    results = {}
    xvars = ['has_sister','gender','par_edu','hukou']
    subjects = [('gpa','GPA'), ('chinese_std','语文'),
                ('math_std','数学'), ('english_std','英语')]
    rows = []
    for fam, fname in [(1,'二孩家庭'), (0,'多孩家庭')]:
        col = 'is_two_child' if fam == 1 else 'is_multi'
        sub = samp_B[samp_B[col] == 1].dropna(subset=xvars+['gpa'])
        for y, label in subjects:
            m = ols(sub, y, xvars)
            r = extract_row(m, 'has_sister', label=f'{fname}_{label}')
            r['家庭类型'] = fname
            r['科目'] = label
            r['样本量'] = int(m.nobs)
            r['R²'] = round(m.rsquared, 4)
            rows.append(r)
    results['表7_分样本异质性'] = pd.DataFrame(rows)
    print("✅ 表7 完成")
    return results


# ── 5.6 家庭规模异质性：交互项回归（表8）────────────────────────────────────

def analysis_family_size_interact(samp_B):
    """
    表8：has_sister × is_multi 交互项（全样本，四科目）
    同时输出全模型系数表
    """
    results = {}
    xvars_base = ['has_sister','is_multi','gender','par_edu','hukou']
    subjects = [('gpa','GPA'), ('chinese_std','语文'),
                ('math_std','数学'), ('english_std','英语')]

    interact_rows = []
    full_models   = {}

    for y, label in subjects:
        samp = samp_B.copy()
        samp['interact'] = samp['has_sister'] * samp['is_multi']
        m = ols(samp, y, xvars_base + ['interact'])
        full_models[label] = m
        for v, lbl in [
            ('has_sister',    f'{label}_有姐姐（主效应，二孩基准）'),
            ('is_multi',f'{label}_多孩家庭（主效应）'),
            ('interact',      f'{label}_有姐姐×多孩家庭（交互项）'),
            ('gender',        f'{label}_学生性别'),
            ('par_edu',       f'{label}_父母教育'),
        ]:
            r = extract_row(m, v, label=lbl)
            r['科目'] = label
            interact_rows.append(r)
        interact_rows.append({
            '变量': f'{label}_样本量', '系数': int(m.nobs),
            '科目': label, '标准误':'','p值':'','显著性':'',
            '95%CI下界':'','95%CI上界':'',
        })
        interact_rows.append({
            '变量': f'{label}_R²', '系数': round(m.rsquared, 4),
            '科目': label, '标准误':'','p值':'','显著性':'',
            '95%CI下界':'','95%CI上界':'',
        })

    results['表8_交互项回归'] = pd.DataFrame(interact_rows)
    print("✅ 表8 完成")
    return results, full_models


# ── 5.7 机制分析：有姐姐 vs 有妹妹（表9）────────────────────────────────────

def analysis_mechanism(samp_broad):
    """
    表9：有年长姐姐 vs 有年幼妹妹 同框对比
    检验姐姐效应来源于"年长指导"而非"家庭有女孩"
    注意：此分析使用宽样本（samp_A，全部二孩家庭），而非主样本C。
    原因：主样本C（二孩次子女）中 has_younger_sister 无变异（所有人均为家中最小孩），
    需要在宽样本中纳入有年幼妹妹的被试才能识别出"年长"效应。
    """
    results = {}
    subjects = [('gpa','GPA'), ('chinese_std','语文'),
                ('math_std','数学'), ('english_std','英语')]
    rows = []
    for y, label in subjects:
        m = ols(samp_broad, y,
                ['has_sister','has_younger_sister','gender','par_edu','hukou'])
        for v, lbl in [('has_sister',         f'{label}_有年长姐姐'),
                       ('has_younger_sister',  f'{label}_有年幼妹妹')]:
            r = extract_row(m, v, label=lbl)
            r['科目'] = label
            rows.append(r)
        rows.append({'变量': f'{label}_样本量', '系数': int(m.nobs),
                     '科目':label,'标准误':'','p值':'','显著性':'',
                     '95%CI下界':'','95%CI上界':''})
    results['表9_机制分析（宽样本A）'] = pd.DataFrame(rows)
    print("✅ 表9 完成（机制分析使用宽样本A，以保证 has_younger_sister 变异）")
    return results


# ── 5.8 稳健性检验四合一（表10）─────────────────────────────────────────────

def analysis_robustness(samp_main, samp_B, samp_broad):
    """
    表10：四类稳健性检验（均基于主回归样本C：严格对比样本）
      检验一：认知能力替代因变量（samp_main）
      检验二：剔除极端值±1%（samp_main）
      检验三：Bootstrap置信区间（samp_main，500次重抽样）
      检验四：宽样本灵敏度对比（samp_broad=样本A，有姐姐 vs 无年长姐姐，全二孩家庭）
              ─ 评估在更宽泛对照组下结果是否一致
    """
    results = {}
    rows    = []

    xvars = ['has_sister','gender','par_edu','hukou']

    # ── 检验一：认知能力替代 (samp_main = 样本C) ──────────────────────────────
    if samp_main['cognitive'].notna().sum() > 50:
        m1 = ols(samp_main, 'cognitive', xvars)
        r  = extract_row(m1, 'has_sister', label='检验一_有姐姐→认知能力（样本C）')
        r.update({'检验': '检验一：认知能力替代（样本C）',
                  '样本量': int(m1.nobs), 'R²': round(m1.rsquared,4)})
        rows.append(r)
        results['表10_检验一_完整系数'] = model_summary(m1, xvars)
    else:
        print("  ⚠ 检验一：认知能力变量缺失或样本不足，跳过")

    # ── 检验二：剔除极端值±1%（samp_main = 样本C）────────────────────────────
    p1, p99 = samp_main['gpa'].quantile([0.01, 0.99])
    samp_trim = samp_main[(samp_main['gpa'] >= p1) & (samp_main['gpa'] <= p99)].copy()
    m2 = ols(samp_trim, 'gpa', xvars)
    r2 = extract_row(m2, 'has_sister', label='检验二_有姐姐（剔除±1%极端值，样本C）')
    r2.update({'检验': '检验二：剔除极端值（样本C）',
               '样本量': int(m2.nobs), 'R²': round(m2.rsquared,4)})
    rows.append(r2)
    results['表10_检验二_完整系数'] = model_summary(m2, xvars)

    # ── 检验三：Bootstrap置信区间（samp_main = 样本C）────────────────────────
    print("  Bootstrap 主分析中（样本C，500次）…")
    beta_bs, se_bs, p_bs, ci_lo, ci_hi = bootstrap_coef(
        samp_main, 'gpa', xvars, 'has_sister', n_boot=500)
    sig_bs = '***' if p_bs<0.01 else ('**' if p_bs<0.05 else ('*' if p_bs<0.10 else ''))
    rows.append({
        '变量': '检验三_Bootstrap（样本C）',
        '系数': round(beta_bs,4), '标准误': round(se_bs,4),
        'p值': round(p_bs,4), '显著性': sig_bs,
        '95%CI下界': round(ci_lo,4), '95%CI上界': round(ci_hi,4),
        '检验': '检验三：Bootstrap（样本C）', '样本量': int(samp_main['gpa'].notna().sum()),
    })
    bs_summary = pd.DataFrame([{
        '系数名': 'has_sister（样本C，Bootstrap）',
        'Bootstrap均值': round(beta_bs,4), 'Bootstrap SE': round(se_bs,4),
        'p值': round(p_bs,4), '95%CI下界': round(ci_lo,4), '95%CI上界': round(ci_hi,4),
    }])
    results['表10_Bootstrap（样本C）'] = bs_summary

    # ── 检验四：宽样本灵敏度对比（samp_broad = 样本A，全二孩家庭） ─────────────
    # 在更宽泛的对照组（所有无年长姐姐的二孩次子女，含有哥哥和无任何年长同胞）下重跑
    m4 = ols(samp_broad, 'gpa', xvars)
    r4 = extract_row(m4, 'has_sister', label='检验四_宽样本（样本A：有姐 vs 全部无姐）')
    r4.update({'检验': '检验四：宽样本灵敏度（样本A）',
               '样本量': int(m4.nobs), 'R²': round(m4.rsquared,4)})
    rows.append(r4)
    results['表10_检验四_完整系数（样本A）'] = model_summary(m4, xvars)

    # 多孩严格对比（samp_B 内部严格对比，附加参考）
    if 'is_multi' in samp_B.columns:
        samp_C_multi = samp_B[
            (samp_B['is_multi']==1) &
            ((samp_B['b0203']>=1)|(samp_B['b0201']>=1)) &
            (samp_B['is_secondborn']==1)
        ].copy()
        if len(samp_C_multi) >= 50:
            xvar_labels = {
                'has_sister': '有姐姐（年长）',
                'gender':     '学生性别',
                'par_edu':    '父母平均教育年限',
                'hukou':      '城镇户口',
                'const':      '常数项',
            }
            subjects_multi = [
                ('chinese_std', '语文'),
                ('math_std',    '数学'),
                ('english_std', '英语'),
                ('gpa',         'GPA'),
            ]

            # ── 参考行（汇总表用）────────────────────────────────────────────────
            m4m = ols(samp_C_multi, 'gpa', xvars)
            r4m = extract_row(m4m, 'has_sister', label='参考_多孩严格对比（有姐 vs 有哥）')
            r4m.update({'检验': '参考：多孩严格对比',
                        '样本量': int(m4m.nobs), 'R²': round(m4m.rsquared,4)})
            rows.append(r4m)

            # ── 完整宽格式（行=变量，列=科目）供表12并列回归使用 ─────────────────
            models_multi = {label: ols(samp_C_multi, y, xvars)
                            for y, label in subjects_multi}

            wide_rows_m = []
            for v in xvars + ['const']:
                row = {'变量': xvar_labels.get(v, v)}
                for _, label in subjects_multi:
                    m = models_multi[label]
                    if v in m.params.index:
                        b   = m.params[v]
                        se  = m.bse[v]
                        p   = m.pvalues[v]
                        sig = '***' if p<0.01 else ('**' if p<0.05 else ('*' if p<0.10 else ''))
                        row[label] = f'{b:.4f}{sig} ({se:.4f})'
                    else:
                        row[label] = '—'
                wide_rows_m.append(row)
            for stat, getter in [('样本量', lambda m: int(m.nobs)),
                                ('R²',    lambda m: round(m.rsquared, 4)),
                                ('调整R²',lambda m: round(m.rsquared_adj, 4))]:
                row = {'变量': stat}
                for _, label in subjects_multi:
                    row[label] = getter(models_multi[label])
                wide_rows_m.append(row)
            results['表12_多孩严格对比_完整宽格式'] = pd.DataFrame(wide_rows_m)

            # ── 数值明细（β / SE / p 分行，便于核对）────────────────────────────
            num_rows_m = []
            for v in xvars + ['const']:
                for stat, col_sfx in [('系数',''), ('标准误',''), ('p值','')]:
                    row = {'变量': xvar_labels.get(v, v), '统计量': stat}
                    for _, label in subjects_multi:
                        m = models_multi[label]
                        if v in m.params.index:
                            if stat == '系数':   row[label] = round(m.params[v], 4)
                            elif stat == '标准误': row[label] = round(m.bse[v], 4)
                            else:                row[label] = round(m.pvalues[v], 4)
                        else:
                            row[label] = '—'
                    num_rows_m.append(row)
            for stat, getter in [('样本量', lambda m: int(m.nobs)),
                                ('R²',    lambda m: round(m.rsquared, 4))]:
                row = {'变量': stat, '统计量': ''}
                for _, label in subjects_multi:
                    row[label] = getter(models_multi[label])
                num_rows_m.append(row)
            results['表12_多孩严格对比_数值明细'] = pd.DataFrame(num_rows_m)
            print(f"  多孩严格对比完整回归已输出（N={int(m4m.nobs):,}）")

    results['表10_四类稳健性汇总'] = pd.DataFrame(rows)

    # ── 普通标准误对比（样本C，OLS vs HC3）──────────────────────────────────
    m_ols = ols(samp_main, 'gpa', xvars, robust='nonrobust')
    results['表10_普通标准误（样本C）'] = model_summary(m_ols, xvars)

    print("✅ 表10 完成（稳健性检验均基于严格对比主样本C）")
    return results, beta_bs, p_bs, ci_lo, ci_hi


# ── 5.9 补充分析：户籍与父母教育异质性 ──────────────────────────────────────

def analysis_extra_heterogeneity(samp_main, samp_B):
    """分城乡、分父母教育程度的异质性分析（主回归样本C：严格对比）"""
    results = {}
    xvars = ['has_sister','gender','par_edu','hukou']
    rows = []

    for label, samp, cond_name in [
        ('城镇', samp_main[samp_main['hukou']==1], '城镇户口（样本C）'),
        ('农村', samp_main[samp_main['hukou']==0], '农村户口（样本C）'),
        ('父母教育高', samp_main[samp_main['par_edu'] >= samp_main['par_edu'].median()], '父母教育≥中位数（样本C）'),
        ('父母教育低', samp_main[samp_main['par_edu'] <  samp_main['par_edu'].median()], '父母教育<中位数（样本C）'),
    ]:
        m = ols(samp, 'gpa', ['has_sister','gender','par_edu','hukou'])
        r = extract_row(m, 'has_sister', label=cond_name)
        r.update({'分组': label, '样本量': int(m.nobs)})
        rows.append(r)

    results['附加_异质性分析（样本C）'] = pd.DataFrame(rows)
    print("✅ 附加异质性分析完成（主回归样本：严格对比样本C）")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# 6. Excel 输出
# ─────────────────────────────────────────────────────────────────────────────

def write_excel(all_results: dict, path: str):
    """将所有分析结果写入一个多 Sheet 的 xlsx 文件，带格式"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空 sheet

    hdr_fill = PatternFill("solid", fgColor="2B3A52")
    hdr_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    thin     = Side(style='thin', color='DDDDDD')
    cell_brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aln = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    for sheet_name, df in all_results.items():
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        if isinstance(df, pd.DataFrame):
            # 写表头
            for ci, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = center
            # 写数据
            for ri, row in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci,
                                   value=val if not (isinstance(val, float) and np.isnan(val)) else '')
                    cell.font = body_font
                    cell.border = cell_brd
                    cell.alignment = center if ci > 1 else left_aln
                    # 斑马纹
                    if ri % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="F5F7FA")
            # 自适应列宽
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)
    print(f"\n✅ Excel 已保存：{path}")
    print(f"   包含 {len(wb.sheetnames)} 个工作表：")
    for s in wb.sheetnames:
        print(f"   • {s}")


# ─────────────────────────────────────────────────────────────────────────────
# 7. 图表生成（6张论文图）
# ─────────────────────────────────────────────────────────────────────────────

def gen_figures(samp_main, samp_B, samp_broad):
    """
    根据实际回归结果生成六张图
    samp_main  = 主回归样本（样本C：严格对比，有年长姐姐 vs 有年长哥哥）
    samp_B     = 全部非独生子女（家庭规模异质性）
    samp_broad = 宽二孩样本（样本A，灵敏度对比）
    """

    # ── 汇总关键系数（从实际回归取值，均基于主样本C）────────────────────────
    xvars = ['has_sister','gender','par_edu','hukou']
    m_gpa = ols(samp_main, 'gpa', xvars)
    m_cn  = ols(samp_main, 'chinese_std', xvars)
    m_mt  = ols(samp_main, 'math_std',    xvars)
    m_en  = ols(samp_main, 'english_std', xvars)

    coef  = [m_gpa.params['has_sister'],  m_cn.params['has_sister'],
             m_mt.params['has_sister'],   m_en.params['has_sister']]
    se    = [m_gpa.bse['has_sister'],    m_cn.bse['has_sister'],
             m_mt.bse['has_sister'],     m_en.bse['has_sister']]
    pvals = [m_gpa.pvalues['has_sister'], m_cn.pvalues['has_sister'],
             m_mt.pvalues['has_sister'],  m_en.pvalues['has_sister']]
    labels = ['GPA', '语文', '数学', '英语']

    def sig_star(p):
        return '***' if p<0.01 else ('**' if p<0.05 else ('†' if p<0.10 else ''))




# ─────────────────────────────────────────────────────────────────────────────
# 8. 主流程
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print(" CEPS 姐姐效应完整分析")
    print(" 主回归样本：样本C（严格对比：二孩次子女，有年长姐姐 vs 有年长哥哥）")
    print("=" * 70)

    # ① 加载 & 构建变量
    raw    = load_data()          # 四文件自动合并
#    diagnose_columns(raw)         # 首次运行：打印关键列名，确认 COL_MAP 后注释掉此行
    df     = build_variables(raw)

    # ② 构建三套样本
    # samp_A = 宽二孩样本（机制分析、灵敏度对比用）
    # samp_B = 全部非独生子女（家庭规模异质性用）
    # samp_C = 严格对比主样本（有年长姐姐 vs 有年长哥哥的二孩次子女）← 主分析
    # samp_C_multi = 多孩严格对比（有年长姐姐 vs 有年长哥哥的多孩次子女）← 表12多孩行
    samp_A, samp_B, samp_C, samp_C_multi = build_samples(df)

    # ③ 各节分析（samp_C 为主样本，samp_A 为辅助）
    print("\n── 描述性统计 ──")
    res_desc = analysis_descriptive(samp_C, samp_B, samp_broad=samp_A,
                                    samp_C_multi=samp_C_multi)

    print("\n── 主回归（表4）── [主样本：严格对比样本C]")
    res_main = analysis_main_regression(samp_C)

    print("\n── 分科目（表5）── [主样本：严格对比样本C]")
    res_subj, _ = analysis_subjects(samp_C)

    print("\n── 性别异质性（表6）── [主样本：严格对比样本C]")
    res_gend = analysis_gender(samp_C)

    print("\n── 分样本异质性（表7）── [样本B：家庭规模]")
    res_fam7 = analysis_family_size_subsample(samp_B)

    print("\n── 交互项（表8）── [样本B：家庭规模]")
    res_fam8, _ = analysis_family_size_interact(samp_B)

    print("\n── 机制分析（表9）── [宽样本A：需要 has_younger_sister 变异]")
    res_mech = analysis_mechanism(samp_A)

    print("\n── 稳健性检验（表10）── [主样本：样本C；灵敏度：样本A]")
    res_rob, beta_bs, p_bs, ci_lo, ci_hi = analysis_robustness(samp_C, samp_B, samp_A)

    print("\n── 附加异质性 ── [主样本：严格对比样本C]")
    res_extra = analysis_extra_heterogeneity(samp_C, samp_B)

    # ④ 汇总 → Excel
    all_results = {}
    all_results.update(res_desc)
    all_results.update(res_main)
    all_results.update(res_subj)
    all_results.update(res_gend)
    all_results.update(res_fam7)
    all_results.update(res_fam8)
    all_results.update(res_mech)
    all_results.update(res_rob)
    all_results.update(res_extra)

    write_excel(all_results, 'CEPS_完整分析结果2.xlsx')

    # ⑤ 生成图表（主样本：样本C；宽样本：样本A）
    print("\n── 生成图表 ──")
    gen_figures(samp_C, samp_B, samp_A)

    print("\n" + "=" * 70)
    print(" 全部分析完成！")
    print("  主回归样本：样本C（严格对比，N = 二孩次子女严格对比组）")
    print("  新增输出表：表5_分科目_完整宽格式  ← 论文表11直接使用")
    print("             表5_分科目_数值明细    ← 所有系数/SE/p值")
    print("             表10_二孩均值差t检验   ← 表10显著性星号依据")
    print("             表12_多孩严格对比描述统计 ← 表12多孩行数据")
    print("             表12_多孩均值差t检验")
    print("  输出文件：CEPS_完整分析结果2.xlsx")

    print("=" * 70)


if __name__ == '__main__':
    main()